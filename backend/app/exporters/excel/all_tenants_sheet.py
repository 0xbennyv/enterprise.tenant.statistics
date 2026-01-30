# app/exporters/excel/all_tenants_sheet.py

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def build_all_tenants_sheet(
    wb: Workbook,
    alerts,
    sla,
    mttd,
    mtta,
    mttr,
    endpoint,
):
    ws = wb.create_sheet("All Tenants")
    # print("Mttd:", mttd)

    

    # ALERTS
    ws.append(["Number of Security Incidents"])

    current_row = 2  # Start after title

    def add_horizontal_section(ws, label, data, start_row):
        """
        label: str -> section name, goes in column A
        data: dict -> keys and values to show horizontally
        start_row: int -> row to start writing
        """
        # Section label in column A of the first row
        ws.cell(row=start_row, column=1, value=label)
        
        # Keys in columns B onward (same row as label)
        for col_idx, key in enumerate(data.keys(), start=2):
            ws.cell(row=start_row, column=col_idx, value=key.title() if label=="Incidents per Severity" else key)
        
        # Values row (row below keys)
        values_row = start_row + 1
        for col_idx, value in enumerate(data.values(), start=2):
            ws.cell(row=values_row, column=col_idx, value=value)
        
        # Return the next row to start for next section
        return values_row + 1

    # Add sections
    current_row = add_horizontal_section(ws, "Incidents per Severity", alerts["total_incident_severity"], current_row)
    current_row = add_horizontal_section(ws, "Incidents per Category", alerts["total_incident_category"], current_row)
    current_row = add_horizontal_section(ws, "Incidents per Month", alerts["total_tenant_monthly"], current_row)
    # END ALERTS

    # SLA
    ws.append([])
    ws.append(["SLA Metrics"])
    for k, v in sla["total_incident_sla_metrics"].items():
        ws.append([k, v])
    ws.append(["Total Incidents", sla["total_incident_count"]])
    # END SLA

    # MTTD
    ws.append([])
    ws.append(["Mean Time to Detect (seconds)", mttd["all_tenants_mttd_seconds"]])
    ws.append(["Total Detections", mttd["total_detections"]])

    # MTTA
    ws.append([])
    ws.append(["Mean Time to Acknowledge (seconds)", mtta["mtta_seconds"]])
    ws.append(["Total Detections", mtta["total_cases"]])

    # MTTR
    ws.append([])
    ws.append(["Mean Time to Recover (seconds)", mttr["mttr_seconds"]])
    ws.append(["Cases", mttr["total_cases"]])

    # ENDPOINT HEALTH
    ws.append([])
    ws.append(["Endpoints Not Fully Protected"])
    # ws.append([[], "Not Fully Protected", "Tamper Disabled"])
    # ws.append("Computer", "Server")

    global_data = endpoint["global"]

    ws.append([
        "Not Fully Protected",
        global_data["notFullyProtected"]])
    ws.append([
        "Tamper Protection Disabled",
        global_data["tamperProtectionDisabled"],
    ])

    # Optional: bold the section labels
    # for row in ws.iter_rows(min_row=2, max_row=30, min_col=1, max_col=1):
    #     for cell in row:
    #         cell.font = Font(bold=True)
    #         cell.alignment = Alignment(vertical="center")
    for cell in ws["A"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name, e.g., 'A'
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2  # extra padding
