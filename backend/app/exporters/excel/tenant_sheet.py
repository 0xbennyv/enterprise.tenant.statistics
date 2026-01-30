# app/exporters/excel/tenant_sheet.py

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from app.utils.helper import unique_sheet_title

def build_tenant_sheet(
    wb: Workbook,
    tenant,
    alerts,
    sla,
    mttd,
    mtta,
    mttr,
    endpoint,
):
    ws = wb.create_sheet(
        unique_sheet_title(wb, tenant["tenantName"])
    )

    # ws.append(["Tenant", tenant["tenantName"]])
    # ws.append([])

    tenant_alerts = next(
        i for i in alerts["incidents"] if i["tenantId"] == tenant["tenantId"]
    )
    # write_incidents_sheet(ws, tenant_alerts)

    current_row = 1  # Start from row 1
    ws.append(["Number of Security Incidents"])
    # Title
    # ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    # ws.cell(row=current_row, column=1, value="Number of Security Incidents")
    # ws["A1"].font = Font(bold=True)
    # ws["A1"].alignment = Alignment(horizontal="center")
    # current_row += 1  # Next row

    def add_horizontal_section(ws, label, data, start_row):
        """
        label: str -> section name, goes in column A
        data: dict -> keys and values to show horizontally
        start_row: int -> row to start writing
        """
        # Section label in column A of the first row
        ws.cell(row=start_row, column=1, value=label)
        
        # Keys in columns B onward (same row as label)
        for col_idx, key in enumerate(data.keys(), start=2):
            ws.cell(row=start_row, column=col_idx, value=key.title() if label=="Incidents per Severity" else key)
        
        # Values row (row below keys)
        values_row = start_row + 1
        for col_idx, value in enumerate(data.values(), start=2):
            ws.cell(row=values_row, column=col_idx, value=value)
        
        # Return the next row to start for next section
        return values_row + 1

    # Add sections
    current_row = add_horizontal_section(ws, "Incidents per Severity", tenant_alerts["severity"], current_row)
    current_row = add_horizontal_section(ws, "Incidents per Category", tenant_alerts["category"], current_row)
    current_row = add_horizontal_section(ws, "Incidents per Month", tenant_alerts["monthly"], current_row)
    # END ALERTS

    tenant_sla = next(
        i for i in sla["incidents"] if i["tenantId"] == tenant["tenantId"]
    )

    # SLA
    ws.append([])
    ws.append(["SLA Metrics"])
    for k, v in tenant_sla["sla_metrics"].items():
        ws.append([k, v])
    ws.append(["Total Incidents", tenant_sla["total_incidents"]])
    # END SLA

    mttd_cases = next(
        i for i in mttd["incidents"] if i["tenantId"] == tenant["tenantId"]
    )
    ws.append([])
    ws.append(["Mean Time to Detect (seconds)", mttd_cases["mttd_seconds"]])
    ws.append(["Total Detections", mttd_cases["total_detections"]])

    mtta_cases = next(
        i for i in mtta["incidents"] if i["tenantId"] == tenant["tenantId"]
    )
    ws.append([])
    ws.append(["Mean Time to Acknowledge (seconds)", mtta_cases["mtta_seconds"]])
    ws.append(["Total Detections", mtta_cases["total_cases"]])

    mttr_cases = next(
        i for i in mttr["incidents"] if i["tenantId"] == tenant["tenantId"]
    )
    ws.append([])
    ws.append(["Mean Time to Recover (seconds)", mttr_cases["mttr_seconds"]])
    ws.append(["Cases", mttr_cases["total_cases"]])

    ws.append([])
    endpoint_data = next(
        i for i in endpoint["tenants"] if i["tenantId"] == tenant["tenantId"]
    )
    # ws.append(["Endpoints Not Protected"])

    # ep = endpoint["by_tenant"][tenant["tenantId"]]
    # ws.append([])
    ws.append(["Endpoints Not Protected"])
    ws.append(["", "Not Fully Protected", "Tamper Disabled"])
    # ws.append(["Computer", ])
    for detail in endpoint_data.get("details", []):
        ws.append([
            detail.get("type"),
            detail.get("notFullyProtected", 0),
            detail.get("tamperProtectionDisabled", 0),
        ])
        # ws.append([
        #     "TOTAL",
        #     endpoint_data.get("notFullyProtected", 0),
        #     endpoint_data.get("tamperProtectionDisabled", 0),
        # ])
    
    # Set Col A to bold
    for cell in ws["A"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    # Set column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name, e.g., 'A'
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2  # extra padding


def write_incidents_sheet(ws, tenant_alerts):
    # Title
    ws.append(["Number of Security Incidents"])
    ws["A1"].font = Font(bold=True)

    # Helper function for each section
    def add_section(ws, title, data, start_row):
        # Section label in column B
        ws.cell(row=start_row, column=2, value=title)
        
        # # Keys in the row below the label
        # keys_row = start_row + 1
        # for col_idx, key in enumerate(data.keys(), start=3):
        #     ws.cell(row=keys_row, column=col_idx, value=key)
        
        # # Values row (below keys)
        # values_row = keys_row + 1
        # for col_idx, value in enumerate(data.values(), start=3):
        #     ws.cell(row=values_row, column=col_idx, value=value)
        
        # Return next start row (2 rows below values row)
        # return values_row + 1

        for k, v in data.items():
            ws.append([k.title(), v])

        return start_row + len(data) + 2  # +2 for title and spacing

    current_row = 2  # Start after title
    current_row = add_section(ws, "Incidents per Severity", tenant_alerts["severity"], current_row)
    current_row = add_section(ws, "Incidents per Category", tenant_alerts["category"], current_row)
    current_row = add_section(ws, "Incidents per Month", tenant_alerts["monthly"], current_row)